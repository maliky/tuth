from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from django.contrib.auth.models import User
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook
from tablib import Dataset

from app.academics.admin import CurriculumCourseResource, CourseWidget
from app.academics.models import College, Course
from app.people.models import RoleAssignment
from app.shared.constants import DEFAULT_ROLE_TO_COLLEGE, TEST_PW
from app.spaces.admin import RoomWidget
from app.spaces.models import Room
from app.timetable.admin import SemesterWidget, SectionWidget
from app.timetable.models import Section, Semester

SECTION_RE = re.compile(r"(?P<sem>\d{2}-\d{2}_Sem\d+):s(?P<num>\d+)")


class Command(BaseCommand):
    """Import data from an Excel workbook."""

    help = "Load workbook data into the database."

    def add_arguments(self, parser):
        parser.add_argument("file", type=str, help="Path to workbook")
        parser.add_argument(
            "--dry-run", action="store_true", default=False, help="Validate only"
        )

    @transaction.atomic
    def handle(self, *args, **opts):
        path = Path(opts["file"])
        if not path.exists():
            raise CommandError(f"File not found: {path}")

        dry_run: bool = opts["dry_run"]
        wb = load_workbook(path, read_only=True, data_only=True)

        totals = {"sections": 0, "courses": 0, "users": 0, "rooms": 0, "errors": 0}
        if "timetable" in wb.sheetnames:
            totals["sections"] = self._import_timetable(wb["timetable"], dry_run)
        if "academics" in wb.sheetnames:
            totals["courses"] = self._import_academics(wb["academics"], dry_run)
        if "people" in wb.sheetnames:
            totals["users"] = self._import_people(wb["people"], dry_run)
        if "spaces" in wb.sheetnames:
            totals["rooms"] = self._import_spaces(wb["spaces"], dry_run)

        msg = (
            f"{totals['sections']} sections, {totals['courses']} courses, "
            f"{totals['users']} users, {totals['rooms']} rooms processed."
        )
        self.stdout.write(self.style.SUCCESS(msg))

    # ------------------------------------------------------------------ helpers
    def _headers(self, ws) -> list[str]:
        return [str(c.value).strip().lower() for c in next(ws.iter_rows(max_row=1))]

    def _as_dict(self, headers: list[str], row: tuple[Any, ...]) -> dict[str, Any]:
        return {h: (row[i] if i < len(row) else None) for i, h in enumerate(headers)}

    def _import_timetable(self, ws, dry_run: bool) -> int:
        headers = self._headers(ws)
        cw = CourseWidget(model=Course, field="code")
        sw = SemesterWidget(model=Semester, field="id")
        scw = SectionCodeWidget(model=Section, field="id")
        created = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            data = self._as_dict(headers, row)
            text = str(data.get("section") or "")
            m = SECTION_RE.match(text)
            if not m:
                continue
            sem_token = m.group("sem")
            sec_no = int(m.group("num"))
            try:
                semester = sw.clean(sem_token, data)
                course_token = data.get("course") or data.get("course_code")
                course = cw.clean(course_token, data)
                if dry_run:
                    continue
                Section.objects.get_or_create(
                    course=course, semester=semester, number=sec_no
                )
                created += 1
            except Exception:
                pass
        return created

    def _import_academics(self, ws, dry_run: bool) -> int:
        headers = self._headers(ws)
        cw = CourseWidget(model=Course, field="code")
        cc_resource = CurriculumCourseResource()
        cc_data = Dataset(headers=["curriculum_name", "college", "course"])
        created = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            data = self._as_dict(headers, row)
            token = data.get("course") or data.get("code")
            if not token:
                continue
            course = cw.clean(token, data)
            title = data.get("title") or course.title
            credit = data.get("credit_hours") or course.credit_hours
            if not dry_run:
                course.title = title
                if credit:
                    course.credit_hours = int(credit)
                course.save()
            created += 1
            cur = data.get("curriculum")
            if cur:
                cc_data.append([cur, data.get("college"), token])
        if not dry_run and len(cc_data):
            cc_resource.import_data(cc_data, raise_errors=False, dry_run=False)
        return created

    def _import_people(self, ws, dry_run: bool) -> int:
        headers = self._headers(ws)
        created = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            data = self._as_dict(headers, row)
            initials = str(data.get("initials") or "").lower()
            lastname = str(data.get("lastname") or "").lower()
            if not initials or not lastname:
                continue
            username = f"{initials}.{lastname}"
            role = str(data.get("role") or "")
            college_code = data.get("college") or DEFAULT_ROLE_TO_COLLEGE.get(role)
            college = None
            if college_code:
                college, _ = College.objects.get_or_create(
                    code=college_code, defaults={"fullname": college_code}
                )
            if dry_run:
                continue
            user, _ = User.objects.get_or_create(
                username=username,
                defaults={
                    "password": TEST_PW,
                    "first_name": data.get("firstname", ""),
                    "last_name": lastname,
                },
            )
            RoleAssignment.objects.get_or_create(
                user=user,
                role=role,
                college=college,
                defaults={"start_date": timezone.now().date()},
            )
            created += 1
        return created

    def _import_spaces(self, ws, dry_run: bool) -> int:
        headers = self._headers(ws)
        rw = RoomWidget(model=Room, field="name")
        created = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            data = self._as_dict(headers, row)
            location = data.get("location")
            if not location:
                continue
            room = rw.clean(location, data)
            if not dry_run and room:
                room.save()
                created += 1
        return created
