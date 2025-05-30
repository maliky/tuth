from __future__ import annotations

from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from import_export import resources
from tablib import Dataset

from app.academics.admin.resources import CourseResource, CollegeResource
from app.timetable.admin.resources import (
    AcademicYearResource,
    SectionResource,
    SemesterResource,
)
from app.people.models import StudentProfile
from app.registry.models import Registration


class StudentResource(resources.ModelResource):
    """Resource for bulk importing :class:`StudentProfile` rows."""

    class Meta:
        model = StudentProfile
        import_id_fields = ("student_id",)
        fields = (
            "student_id",
            "user",
            "college",
            "curriculum",
            "enrollment_semester",
            "enrollment_date",
        )


class RegistrationResource(resources.ModelResource):
    """Resource for bulk importing :class:`Registration` rows."""

    class Meta:
        model = Registration
        import_id_fields = ("student", "section")
        fields = (
            "student",
            "section",
            "status",
            "date_latest_reservation",
        )


class Command(BaseCommand):
    """Import TU data from an Excel workbook."""

    help = "Load Tubman University data from an Excel file"

    def add_arguments(self, parser):
        parser.add_argument("file", type=str, help="Path to TU_DB workbook")
        parser.add_argument(
            "--dry-run",
            action="store_true",
            default=False,
            help="Validate data without saving",
        )

    def handle(self, *args, **options):
        path = Path(options["file"])
        if not path.exists():
            raise CommandError(f"File not found: {path}")

        try:
            import openpyxl  # type: ignore
        except Exception as exc:  # pragma: no cover - optional dependency
            raise CommandError(f"openpyxl required: {exc}")

        wb = openpyxl.load_workbook(path)
        resources_map = {
            "Colleges": CollegeResource,
            "Courses": CourseResource,
            "AcademicYears": AcademicYearResource,
            "Semesters": SemesterResource,
            "Sections": SectionResource,
            "Students": StudentResource,
            "Registrations": RegistrationResource,
        }
        for sheet, resource_cls in resources_map.items():
            if sheet not in wb.sheetnames:
                continue
            ws = wb[sheet]
            dataset = Dataset()
            dataset.headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
            for row in ws.iter_rows(min_row=2, values_only=True):
                dataset.append(list(row))

            resource = resource_cls()
            result = resource.import_data(
                dataset, dry_run=options["dry_run"], raise_errors=True
            )
            self.stdout.write(
                self.style.SUCCESS(
                    f"{sheet}: {result.totals.get('new', 0)} created, {result.totals.get('update', 0)} updated"
                )
            )
        self.stdout.write(self.style.SUCCESS("Import complete"))
